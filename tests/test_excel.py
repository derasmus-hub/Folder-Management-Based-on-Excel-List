"""
Unit tests for the Excel CaseID loader.
"""

import tempfile
from pathlib import Path

import openpyxl
import pytest

from folder_mover.excel import load_case_ids


def create_test_xlsx(data: list, sheet_name: str = "Sheet1") -> Path:
    """
    Create a temporary XLSX file with data in Column A.

    Args:
        data: List of values to put in Column A (one per row)
        sheet_name: Name for the worksheet

    Returns:
        Path to the temporary XLSX file
    """
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name

    for row_idx, value in enumerate(data, start=1):
        worksheet.cell(row=row_idx, column=1, value=value)

    # Save to temp file
    temp_file = tempfile.NamedTemporaryFile(
        suffix=".xlsx",
        delete=False
    )
    temp_path = Path(temp_file.name)
    temp_file.close()
    workbook.save(temp_path)
    workbook.close()

    return temp_path


class TestLeadingZeros:
    """Tests for leading zero preservation."""

    def test_leading_zeros_preserved_as_string(self):
        """CaseIDs with leading zeros stored as text are preserved."""
        xlsx_path = create_test_xlsx(["00123", "00456", "00789"])
        try:
            result = load_case_ids(xlsx_path)
            assert result == ["00123", "00456", "00789"]
        finally:
            xlsx_path.unlink()

    def test_numeric_values_converted_to_string(self):
        """Numeric cell values are converted to strings."""
        xlsx_path = create_test_xlsx([123, 456, 789])
        try:
            result = load_case_ids(xlsx_path)
            assert result == ["123", "456", "789"]
            # Verify they are strings, not integers
            assert all(isinstance(cid, str) for cid in result)
        finally:
            xlsx_path.unlink()

    def test_mixed_string_and_numeric(self):
        """Mix of string and numeric values all become strings."""
        xlsx_path = create_test_xlsx(["ABC-001", 12345, "00789", 42])
        try:
            result = load_case_ids(xlsx_path)
            assert result == ["ABC-001", "12345", "00789", "42"]
        finally:
            xlsx_path.unlink()

    def test_float_values_converted(self):
        """Float values are converted to string representation."""
        xlsx_path = create_test_xlsx([123.0, 456.5])
        try:
            result = load_case_ids(xlsx_path)
            # openpyxl may return 123.0 or 123 depending on cell format
            assert len(result) == 2
            assert all(isinstance(cid, str) for cid in result)
        finally:
            xlsx_path.unlink()


class TestEmptyHandling:
    """Tests for empty cell handling."""

    def test_empty_rows_ignored(self):
        """Empty cells are skipped."""
        xlsx_path = create_test_xlsx(["A001", None, "A002", None, "A003"])
        try:
            result = load_case_ids(xlsx_path)
            assert result == ["A001", "A002", "A003"]
        finally:
            xlsx_path.unlink()

    def test_whitespace_only_ignored(self):
        """Cells with only whitespace are skipped."""
        xlsx_path = create_test_xlsx(["A001", "   ", "A002", "\t", "A003"])
        try:
            result = load_case_ids(xlsx_path)
            assert result == ["A001", "A002", "A003"]
        finally:
            xlsx_path.unlink()

    def test_whitespace_trimmed(self):
        """Leading and trailing whitespace is trimmed."""
        xlsx_path = create_test_xlsx(["  A001  ", " A002", "A003 "])
        try:
            result = load_case_ids(xlsx_path)
            assert result == ["A001", "A002", "A003"]
        finally:
            xlsx_path.unlink()

    def test_all_empty_raises_error(self):
        """File with only empty cells raises ValueError."""
        xlsx_path = create_test_xlsx([None, None, None])
        try:
            with pytest.raises(ValueError, match="No CaseIDs found"):
                load_case_ids(xlsx_path)
        finally:
            xlsx_path.unlink()


class TestDeduplication:
    """Tests for duplicate handling."""

    def test_duplicates_removed(self):
        """Duplicate CaseIDs are removed."""
        xlsx_path = create_test_xlsx(["A001", "A002", "A001", "A003", "A002"])
        try:
            result = load_case_ids(xlsx_path)
            assert result == ["A001", "A002", "A003"]
        finally:
            xlsx_path.unlink()

    def test_original_order_preserved(self):
        """First occurrence order is preserved after dedup."""
        xlsx_path = create_test_xlsx(["C", "A", "B", "A", "C", "D"])
        try:
            result = load_case_ids(xlsx_path)
            assert result == ["C", "A", "B", "D"]
        finally:
            xlsx_path.unlink()

    def test_case_sensitive_dedup(self):
        """Deduplication is case-sensitive."""
        xlsx_path = create_test_xlsx(["abc", "ABC", "Abc"])
        try:
            result = load_case_ids(xlsx_path)
            assert result == ["abc", "ABC", "Abc"]
        finally:
            xlsx_path.unlink()


class TestValidation:
    """Tests for input validation."""

    def test_file_not_found(self):
        """Non-existent file raises FileNotFoundError."""
        with pytest.raises(FileNotFoundError, match="not found"):
            load_case_ids("/nonexistent/path/file.xlsx")

    def test_invalid_sheet_name(self):
        """Invalid sheet name raises ValueError."""
        xlsx_path = create_test_xlsx(["A001"])
        try:
            with pytest.raises(ValueError, match="not found"):
                load_case_ids(xlsx_path, sheet_name="NonExistent")
        finally:
            xlsx_path.unlink()

    def test_valid_sheet_name(self):
        """Specifying valid sheet name works."""
        # Create workbook with custom sheet name
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "MyCases"
        worksheet.cell(row=1, column=1, value="CASE001")

        temp_file = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        temp_path = Path(temp_file.name)
        temp_file.close()
        workbook.save(temp_path)
        workbook.close()

        try:
            result = load_case_ids(temp_path, sheet_name="MyCases")
            assert result == ["CASE001"]
        finally:
            temp_path.unlink()


class TestEdgeCases:
    """Tests for edge cases."""

    def test_single_case_id(self):
        """Single CaseID works correctly."""
        xlsx_path = create_test_xlsx(["ONLY-ONE"])
        try:
            result = load_case_ids(xlsx_path)
            assert result == ["ONLY-ONE"]
        finally:
            xlsx_path.unlink()

    def test_special_characters(self):
        """CaseIDs with special characters are preserved."""
        xlsx_path = create_test_xlsx(["A-001", "B_002", "C.003", "D/004"])
        try:
            result = load_case_ids(xlsx_path)
            assert result == ["A-001", "B_002", "C.003", "D/004"]
        finally:
            xlsx_path.unlink()

    def test_unicode_characters(self):
        """Unicode characters in CaseIDs are preserved."""
        xlsx_path = create_test_xlsx(["案件001", "Ñoño", "日本語"])
        try:
            result = load_case_ids(xlsx_path)
            assert result == ["案件001", "Ñoño", "日本語"]
        finally:
            xlsx_path.unlink()

    def test_path_as_string(self):
        """Function accepts string path as well as Path object."""
        xlsx_path = create_test_xlsx(["TEST001"])
        try:
            result = load_case_ids(str(xlsx_path))
            assert result == ["TEST001"]
        finally:
            xlsx_path.unlink()
