"""
Excel file reader for extracting CaseIDs.

This module is responsible for:
- Reading XLSX files using openpyxl
- Extracting CaseIDs from Column A
- Preserving leading zeros by treating all values as strings
- Handling empty cells and invalid data gracefully
- Deduplicating while preserving original order
- Logging warnings for skipped rows
"""

import logging
from pathlib import Path
from typing import List, Union

import openpyxl

logger = logging.getLogger(__name__)


def load_case_ids(
    excel_path: Union[str, Path],
    sheet_name: str = None
) -> List[str]:
    """
    Load CaseIDs from Column A of an Excel file.

    Reads all values from Column A, treating each as a string to preserve
    leading zeros. Empty cells are ignored, whitespace is trimmed, and
    duplicates are removed while preserving the original order.

    Args:
        excel_path: Path to the XLSX file
        sheet_name: Optional sheet name (defaults to active sheet)

    Returns:
        List of unique CaseID strings in original order

    Raises:
        FileNotFoundError: If the Excel file doesn't exist
        ValueError: If the file cannot be parsed or contains no CaseIDs
    """
    path = Path(excel_path)

    # Validate file exists
    if not path.exists():
        raise FileNotFoundError(f"Excel file not found: {path}")

    if not path.is_file():
        raise ValueError(f"Path is not a file: {path}")

    logger.info(f"Loading CaseIDs from: {path}")

    try:
        # Load workbook in read-only mode for better performance
        # data_only=True to get values instead of formulas
        workbook = openpyxl.load_workbook(
            path,
            read_only=True,
            data_only=True
        )
    except Exception as e:
        raise ValueError(f"Failed to open Excel file: {e}") from e

    try:
        # Select worksheet
        if sheet_name:
            if sheet_name not in workbook.sheetnames:
                available = ", ".join(workbook.sheetnames)
                raise ValueError(
                    f"Sheet '{sheet_name}' not found. Available: {available}"
                )
            worksheet = workbook[sheet_name]
            logger.debug(f"Using specified sheet: {sheet_name}")
        else:
            worksheet = workbook.active
            logger.debug(f"Using active sheet: {worksheet.title}")

        # Extract CaseIDs from Column A
        case_ids: List[str] = []
        seen: set = set()
        row_count = 0
        skipped_count = 0
        duplicate_count = 0

        for row in worksheet.iter_rows(min_col=1, max_col=1, values_only=True):
            row_count += 1
            cell_value = row[0]

            # Skip empty cells
            if cell_value is None:
                skipped_count += 1
                logger.debug(f"Row {row_count}: Empty cell, skipping")
                continue

            # Convert to string and trim whitespace
            # This preserves leading zeros for numeric-looking values
            case_id = str(cell_value).strip()

            # Skip empty strings after trimming
            if not case_id:
                skipped_count += 1
                logger.debug(f"Row {row_count}: Empty after trim, skipping")
                continue

            # Deduplicate while preserving order
            if case_id in seen:
                duplicate_count += 1
                logger.debug(f"Row {row_count}: Duplicate '{case_id}', skipping")
                continue

            seen.add(case_id)
            case_ids.append(case_id)
            logger.debug(f"Row {row_count}: Added CaseID '{case_id}'")

    finally:
        workbook.close()

    # Validate we found at least one CaseID
    if not case_ids:
        raise ValueError(
            f"No CaseIDs found in Column A of '{path}'. "
            f"Checked {row_count} rows, all were empty."
        )

    logger.info(
        f"Loaded {len(case_ids)} unique CaseIDs "
        f"(skipped {skipped_count} empty, {duplicate_count} duplicates)"
    )

    return case_ids


# Keep the old function name for backwards compatibility with CLI
def read_case_ids(excel_path: Path, sheet_name: str = None) -> List[str]:
    """
    Read CaseIDs from Column A of an Excel file.

    This is an alias for load_case_ids() for backwards compatibility.

    Args:
        excel_path: Path to the XLSX file
        sheet_name: Optional sheet name (defaults to active sheet)

    Returns:
        List of CaseID strings with leading zeros preserved

    Raises:
        FileNotFoundError: If the Excel file doesn't exist
        ValueError: If the file cannot be parsed
    """
    return load_case_ids(excel_path, sheet_name)
